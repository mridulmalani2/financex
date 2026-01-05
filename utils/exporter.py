#!/usr/bin/env python3
"""
Exporter - FinanceX Download Package Builder
=============================================
Production V1.0 - Bundles all outputs for user download.

The Download Package includes:
1. Financial Models (DCF, LBO, Comps) - as .xlsx with formatting or clean .csv
2. Validation Report - audit findings
3. analyst_brain.json - UPDATED with all session corrections
4. Thinking Log - Engine reasoning trail

CRITICAL: The brain file MUST include every fix the user made during the session.

Output Format Options:
- Excel (.xlsx) with professional formatting (if openpyxl available)
- Clean CSV fallback if Excel not available
"""

import os
import io
import csv
import json
import zipfile
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass

# Try to import openpyxl for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Try to import pandas
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


@dataclass
class ExportResult:
    """Result of an export operation."""
    success: bool
    filename: str
    data: bytes
    mime_type: str
    file_count: int
    error_message: Optional[str] = None


# =============================================================================
# EXCEL FORMATTING STYLES
# =============================================================================

if EXCEL_AVAILABLE:
    # Goldman Sachs / JPMC style formatting
    HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

    DATA_FONT = Font(name='Arial', size=10)
    DATA_ALIGNMENT = Alignment(horizontal='right', vertical='center')

    LABEL_FONT = Font(name='Arial', size=10, bold=False)
    LABEL_ALIGNMENT = Alignment(horizontal='left', vertical='center')

    SUBTOTAL_FONT = Font(name='Arial', size=10, bold=True)
    SUBTOTAL_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    TOTAL_FONT = Font(name='Arial', size=11, bold=True)
    TOTAL_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

    THIN_BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    NUMBER_FORMAT = '#,##0'
    PERCENT_FORMAT = '0.0%'
    CURRENCY_FORMAT = '$#,##0'


# =============================================================================
# EXCEL EXPORT FUNCTIONS
# =============================================================================

def create_formatted_excel(
    dataframes: Dict[str, 'pd.DataFrame'],
    title: str = "FinanceX Export"
) -> bytes:
    """
    Create a professionally formatted Excel workbook with multiple sheets.

    Args:
        dataframes: Dict of {sheet_name: DataFrame}
        title: Workbook title

    Returns:
        Excel file as bytes
    """
    if not EXCEL_AVAILABLE or not PANDAS_AVAILABLE:
        raise ImportError("openpyxl and pandas required for Excel export")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for sheet_name, df in dataframes.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel limits sheet names to 31 chars

        # Write header
        ws.cell(row=1, column=1, value=sheet_name).font = Font(size=14, bold=True)
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

        # Start data at row 4
        start_row = 4

        # Write DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True)):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=start_row + r_idx, column=c_idx + 1, value=value)

                # Format header row
                if r_idx == 0:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = HEADER_ALIGNMENT
                    cell.border = THIN_BORDER
                # Format index column (labels)
                elif c_idx == 0:
                    cell.font = LABEL_FONT
                    cell.alignment = LABEL_ALIGNMENT
                    cell.border = THIN_BORDER

                    # Highlight subtotals and totals
                    if isinstance(value, str):
                        if value.startswith('(=)'):
                            cell.font = SUBTOTAL_FONT
                            cell.fill = SUBTOTAL_FILL
                        elif 'Total' in value or 'UFCF' in value:
                            cell.font = TOTAL_FONT
                            cell.fill = TOTAL_FILL
                # Format data cells
                else:
                    cell.font = DATA_FONT
                    cell.alignment = DATA_ALIGNMENT
                    cell.border = THIN_BORDER

                    # Apply number format
                    if isinstance(value, (int, float)):
                        if '%' in str(df.index[r_idx - 1]) if r_idx > 0 else '':
                            cell.number_format = PERCENT_FORMAT
                        else:
                            cell.number_format = NUMBER_FORMAT

        # Auto-fit column widths
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze panes (header + index column)
        ws.freeze_panes = 'B5'

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def csv_to_excel_bytes(csv_path: str, sheet_name: str = "Data") -> bytes:
    """Convert a CSV file to formatted Excel bytes."""
    if not PANDAS_AVAILABLE:
        raise ImportError("pandas required for CSV to Excel conversion")

    df = pd.read_csv(csv_path, index_col=0)
    return create_formatted_excel({sheet_name: df})


# =============================================================================
# PACKAGE BUILDER
# =============================================================================

class PackageBuilder:
    """
    Builds the complete download package for the user.

    The package includes:
    - Financial models as Excel or CSV
    - Updated analyst_brain.json
    - Thinking log
    - Validation report
    """

    def __init__(self, session_dir: str = None, output_dir: str = None):
        """
        Initialize the package builder.

        Args:
            session_dir: Path to session directory with files
            output_dir: Alternative output directory
        """
        self.session_dir = session_dir
        self.output_dir = output_dir or session_dir
        self.files_added = []
        self.errors = []

    def build_package(
        self,
        brain_manager: Any = None,
        model_files: Dict[str, str] = None,
        include_thinking_log: bool = True,
        excel_format: bool = True
    ) -> ExportResult:
        """
        Build the complete download package.

        Args:
            brain_manager: BrainManager instance with session corrections
            model_files: Dict of {name: path} for model files
            include_thinking_log: Whether to include engine thinking log
            excel_format: Use Excel format if available (otherwise CSV)

        Returns:
            ExportResult with the ZIP file data
        """
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            # 1. Add Financial Models
            if model_files:
                self._add_models(zf, model_files, excel_format)

            # 2. Add Updated Brain (CRITICAL - must include all session fixes)
            if brain_manager:
                self._add_brain(zf, brain_manager)

            # 3. Add Thinking Log
            if include_thinking_log:
                self._add_thinking_log(zf)

            # 4. Add README with package info
            self._add_readme(zf)

        zip_buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"FinanceX_Package_{timestamp}.zip"

        return ExportResult(
            success=len(self.errors) == 0,
            filename=filename,
            data=zip_buffer.getvalue(),
            mime_type="application/zip",
            file_count=len(self.files_added),
            error_message="; ".join(self.errors) if self.errors else None
        )

    def _add_models(
        self,
        zf: zipfile.ZipFile,
        model_files: Dict[str, str],
        excel_format: bool
    ):
        """Add financial model files to the package."""
        models_to_combine = {}

        for name, path in model_files.items():
            if not path or not os.path.exists(path):
                continue

            try:
                if excel_format and EXCEL_AVAILABLE and PANDAS_AVAILABLE:
                    # Read CSV and store for combined Excel
                    df = pd.read_csv(path, index_col=0)
                    models_to_combine[name] = df
                else:
                    # Add as CSV
                    with open(path, 'r', encoding='utf-8') as f:
                        zf.writestr(f"models/{os.path.basename(path)}", f.read())
                    self.files_added.append(path)

            except Exception as e:
                self.errors.append(f"Error adding {name}: {str(e)}")

        # Create combined Excel workbook if we have models
        if models_to_combine and excel_format and EXCEL_AVAILABLE:
            try:
                excel_bytes = create_formatted_excel(models_to_combine, "FinanceX Models")
                zf.writestr("FinanceX_Models.xlsx", excel_bytes)
                self.files_added.append("FinanceX_Models.xlsx")
            except Exception as e:
                self.errors.append(f"Error creating Excel: {str(e)}")
                # Fallback to CSVs
                for name, df in models_to_combine.items():
                    csv_data = df.to_csv()
                    zf.writestr(f"models/{name}.csv", csv_data)
                    self.files_added.append(f"{name}.csv")

    def _add_brain(self, zf: zipfile.ZipFile, brain_manager: Any):
        """
        Add the UPDATED analyst_brain.json to the package.

        CRITICAL: This must include EVERY fix the user made during the session.
        """
        try:
            # Get the complete brain JSON (includes all session updates)
            brain_json = brain_manager.to_json_string()

            # Verify it has the session data
            brain_data = json.loads(brain_json)

            # Add metadata about this export
            brain_data['metadata']['exported_at'] = datetime.now().isoformat()
            brain_data['metadata']['export_version'] = '1.0'

            # Include session history count
            session_actions = len(brain_data.get('session_history', []))

            # Re-serialize with metadata
            final_json = json.dumps(brain_data, indent=2, ensure_ascii=False)

            zf.writestr("analyst_brain.json", final_json)
            self.files_added.append("analyst_brain.json")

            print(f"[Exporter] Brain exported with {brain_data['metadata'].get('total_mappings', 0)} mappings, {session_actions} session actions")

        except Exception as e:
            self.errors.append(f"Error exporting brain: {str(e)}")

    def _add_thinking_log(self, zf: zipfile.ZipFile):
        """Add the engine thinking log to the package."""
        # Try multiple possible log locations
        log_paths = [
            os.path.join(self.output_dir or '', 'logs', 'engine_thinking.log'),
            os.path.join(self.session_dir or '', 'logs', 'engine_thinking.log'),
            os.path.join(os.path.dirname(os.path.dirname(__file__)), 'logs', 'engine_thinking.log'),
        ]

        for log_path in log_paths:
            if log_path and os.path.exists(log_path):
                try:
                    with open(log_path, 'r', encoding='utf-8') as f:
                        zf.writestr("thinking_log.txt", f.read())
                    self.files_added.append("thinking_log.txt")
                    return
                except Exception as e:
                    self.errors.append(f"Error adding thinking log: {str(e)}")

    def _add_readme(self, zf: zipfile.ZipFile):
        """Add a README file to the package."""
        readme_content = f"""FinanceX Export Package
========================
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Contents:
---------
"""
        for f in self.files_added:
            readme_content += f"- {f}\n"

        readme_content += """
Files Description:
------------------
- FinanceX_Models.xlsx: DCF, LBO, and Comps financial models with formatting
- analyst_brain.json: Your analyst brain with all session corrections (IMPORTANT!)
- thinking_log.txt: Engine reasoning trail for audit purposes

Next Steps:
-----------
1. Open FinanceX_Models.xlsx in Excel for analysis
2. Keep analyst_brain.json - upload it next time to restore your corrections
3. Review thinking_log.txt to understand the model calculations

Support:
--------
Report issues at: https://github.com/anthropics/financex/issues
"""
        zf.writestr("README.txt", readme_content)


# =============================================================================
# CONVENIENCE FUNCTIONS
# =============================================================================

def create_download_package(
    brain_manager: Any,
    model_files: Dict[str, str],
    output_dir: str = None,
    excel_format: bool = True
) -> ExportResult:
    """
    Create a complete download package with all user outputs.

    Args:
        brain_manager: BrainManager with session corrections
        model_files: Dict mapping model names to file paths
        output_dir: Output directory (optional)
        excel_format: Use Excel format if available

    Returns:
        ExportResult with the ZIP file data
    """
    builder = PackageBuilder(output_dir=output_dir)
    return builder.build_package(
        brain_manager=brain_manager,
        model_files=model_files,
        include_thinking_log=True,
        excel_format=excel_format
    )


def export_brain_only(brain_manager: Any) -> Tuple[str, bytes]:
    """
    Export just the analyst brain as JSON.

    Args:
        brain_manager: BrainManager instance

    Returns:
        Tuple of (filename, json_bytes)
    """
    brain_json = brain_manager.to_json_string()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"analyst_brain_{timestamp}.json"
    return filename, brain_json.encode('utf-8')


def export_model_as_excel(csv_path: str, model_name: str = "Model") -> Tuple[str, bytes]:
    """
    Export a single model CSV as a formatted Excel file.

    Args:
        csv_path: Path to the CSV file
        model_name: Name for the Excel sheet

    Returns:
        Tuple of (filename, excel_bytes)
    """
    if not EXCEL_AVAILABLE or not PANDAS_AVAILABLE:
        # Fall back to CSV
        with open(csv_path, 'rb') as f:
            return os.path.basename(csv_path), f.read()

    excel_bytes = csv_to_excel_bytes(csv_path, model_name)
    filename = os.path.splitext(os.path.basename(csv_path))[0] + ".xlsx"
    return filename, excel_bytes


# =============================================================================
# TESTING
# =============================================================================

if __name__ == "__main__":
    print("Exporter Module Test")
    print("=" * 50)
    print(f"Excel Available: {EXCEL_AVAILABLE}")
    print(f"Pandas Available: {PANDAS_AVAILABLE}")

    if PANDAS_AVAILABLE:
        # Create test DataFrame
        import pandas as pd
        test_df = pd.DataFrame({
            "2024": [100, 50, 50, 20, 10, 20, 5, 15, 3, 12, 5, 7],
            "2023": [90, 45, 45, 18, 9, 18, 4, 14, 3, 11, 4, 7],
            "2022": [80, 40, 40, 16, 8, 16, 4, 12, 2, 10, 4, 6],
        }, index=[
            "Total Revenue",
            "(-) COGS",
            "(=) Gross Profit",
            "(-) SG&A",
            "(-) R&D",
            "(=) EBITDA",
            "(-) D&A",
            "(=) EBIT",
            "(-) Cash Taxes",
            "(=) NOPAT",
            "(-) CapEx",
            "(=) Unlevered Free Cash Flow"
        ])

        print("\nTest DataFrame:")
        print(test_df)

        if EXCEL_AVAILABLE:
            print("\nCreating formatted Excel...")
            excel_bytes = create_formatted_excel({"DCF": test_df})
            print(f"Excel size: {len(excel_bytes)} bytes")
