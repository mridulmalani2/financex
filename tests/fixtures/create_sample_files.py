#!/usr/bin/env python3
"""
Create sample Excel files for testing.
"""

import os
try:
    from openpyxl import Workbook
except ImportError:
    print("openpyxl not installed, creating placeholder files")
    # Create placeholder files
    for filename in ['clean_company.xlsx', 'missing_revenue.xlsx', 'negative_revenue.xlsx', 'ambiguous_labels.xlsx']:
        filepath = os.path.join(os.path.dirname(__file__), filename)
        with open(filepath, 'w') as f:
            f.write('PLACEHOLDER EXCEL FILE')
    exit(0)

# Create clean_company.xlsx
wb = Workbook()
ws = wb.active
ws.title = "Income Statement"

ws['A1'] = 'Concept'
ws['B1'] = '2024'
ws['A2'] = 'Revenue'
ws['B2'] = 1000000
ws['A3'] = 'EBITDA'
ws['B3'] = 200000
ws['A4'] = 'Net Income'
ws['B4'] = 100000

wb.save(os.path.join(os.path.dirname(__file__), 'clean_company.xlsx'))

# Create missing_revenue.xlsx
wb = Workbook()
ws = wb.active
ws.title = "Income Statement"

ws['A1'] = 'Concept'
ws['B1'] = '2024'
ws['A2'] = 'EBITDA'
ws['B2'] = 200000
ws['A3'] = 'Net Income'
ws['B3'] = 100000

wb.save(os.path.join(os.path.dirname(__file__), 'missing_revenue.xlsx'))

# Create negative_revenue.xlsx
wb = Workbook()
ws = wb.active
ws.title = "Income Statement"

ws['A1'] = 'Concept'
ws['B1'] = '2024'
ws['A2'] = 'Revenue'
ws['B2'] = -1000000  # Negative value
ws['A3'] = 'EBITDA'
ws['B3'] = 200000

wb.save(os.path.join(os.path.dirname(__file__), 'negative_revenue.xlsx'))

# Create ambiguous_labels.xlsx
wb = Workbook()
ws = wb.active
ws.title = "Income Statement"

ws['A1'] = 'Concept'
ws['B1'] = '2024'
ws['A2'] = 'Total Revenues'  # Ambiguous label
ws['B2'] = 1000000
ws['A3'] = 'Operating Profit'  # Ambiguous label
ws['B3'] = 200000

wb.save(os.path.join(os.path.dirname(__file__), 'ambiguous_labels.xlsx'))

print("Created 4 sample Excel files")
